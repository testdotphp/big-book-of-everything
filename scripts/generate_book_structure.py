#!/usr/bin/env python3
"""
Generate book-structure.json from the Big Book of Everything Mk IV Excel file.

Reads /home/teedge/Downloads/bigbookmkIV.xlsx and writes the structured JSON to
/home/teedge/projects/portal-template/configs/book-structure.json
"""

import json
import re
import openpyxl

EXCEL_PATH = "/home/teedge/Downloads/bigbookmkIV.xlsx"
OUTPUT_PATH = "/home/teedge/projects/portal-template/configs/book-structure.json"


def make_slug(name: str) -> str:
    """Lowercase, replace non-alphanumeric with hyphens, strip leading/trailing hyphens."""
    slug = name.lower()
    slug = re.sub(r"[^a-z0-9]+", "-", slug)
    slug = slug.strip("-")
    return slug


def guess_field_type(label: str) -> str:
    """Guess field type from the label text."""
    lower = label.lower().strip()

    # Specific overrides for known edge cases
    OVERRIDES = {
        "who employed?": "text",         # question about who, not a boolean
        "who is the letter for?": "text", # same - question not yes/no
        "paid off (and date)?": "text",   # text field, not boolean
        "paid off and date?": "text",
        "payroll name": "text",           # not currency despite "pay"
        "contact phone/email": "text",    # mixed field
        "date due/how pay": "text",       # mixed field
        "due date/how pay": "text",       # mixed field in insurance sheets
        "how to pay": "text",             # instructions, not currency
        "how pay": "text",               # instructions, not currency
        "automated payments from": "text", # description field
        "website/company": "text",        # company name, not url
        "anniversary": "date",            # keep as date (standalone word)
        # Final Arrangements: open-ended questions, not yes/no booleans
        "where interred?": "text",
        "who performs ceremony?": "text",
        "pallbearers?": "text",
        "special music, notes, food, or drink?": "text",
        "ceremony speakers": "text",
        # Safes & Storage
        "what is stored within?": "text",
        # Estate Plans
        "who should receive this?": "text",
        "what circumstances (death of self, spouse, or both)": "text",
        # Debts
        "original debt amt": "currency",
    }
    if lower in OVERRIDES:
        return OVERRIDES[lower]

    # Boolean: ends with ? (after overrides handled above)
    if lower.endswith("?"):
        return "boolean"

    # Date patterns
    date_keywords = ["date", " dob", "expir", "since", "birthdate", "birthplace date",
                     "adopt/birth date"]
    for kw in date_keywords:
        if kw in lower:
            # Exclude "date due/how pay" - already handled by override
            return "date"

    # "anniversary" as a standalone word
    if "anniversary" in lower:
        return "date"

    # Phone/fax - but not when combined with other things like "phone/email"
    if ("phone" in lower or "fax" in lower) and "email" not in lower:
        return "phone"

    # Email
    if "email" in lower:
        return "email"

    # Website/URL - but not "website/company" (handled by override)
    if "website" in lower or "url" in lower:
        return "url"

    # Currency - more specific patterns to avoid false positives
    # Match on word boundaries where possible
    currency_patterns = [
        r"\bamount\b", r"\bvalue\b", r"\bcost\b", r"\bsalary\b",
        r"\bbalance\b", r"\blimit\b", r"\btax owed\b", r"\btax paid\b",
        r"\bprice\b", r"\bdeduct",
    ]
    for pat in currency_patterns:
        if re.search(pat, lower):
            return "currency"

    # "pay" only when it means payment amount (starting pay, ending pay, co pay amount, payment amount)
    if re.search(r"\bpay\b", lower) and any(kw in lower for kw in ["starting", "ending", "co pay", "payment"]):
        return "currency"

    return "text"


# === Category / Section definitions ===

CATEGORIES = [
    {
        "name": "About Me and My Family",
        "icon": "users",
        "sections": [
            "Personal Info",
            "Adoption Info",
            "Pet Information",
            "Previous Addresses",
            "Current Resume",
            "Employment History",
            "Schooling History",
            "Transcripts",
            "Military History",
            "Groups and Organizations",
            "Credit Report",
        ],
    },
    {
        "name": "Other Information",
        "icon": "file-text",
        "sections": [
            "Document Location",
            "Emergency Plan",
            "Travel Info",
            "Safes & Storage",
            "Tax Issues",
            "Property Tax",
            "Est Tax Payments",
            "Data Backup Plans",
            "Passwords & Logins",
            "Account Numbers",
            "Courses",
        ],
    },
    {
        "name": "Assets",
        "icon": "dollar-sign",
        "sections": [
            "Bank Accounts",
            "Investments",
            "Retirement Plans",
            "Private Business",
            "Real Estate",
            "Vehicles",
            "Valuables Inventory",
            "Debts Owed to You",
            "Other Assets",
        ],
    },
    {
        "name": "Liabilities",
        "icon": "credit-card",
        "sections": [
            "Subscriptions",
            "Loan Obligations",
            "Credit Cards",
            "Other Debts",
        ],
    },
    {
        "name": "Insurance",
        "icon": "shield",
        "sections": [
            "Life Insurance",
            "Health Insurance",
            "Car Insurance",
            "Vehicle Insurance",
            "Home-Renter Insurance",
        ],
    },
    {
        "name": "Medical",
        "icon": "heart-pulse",
        "sections": [
            "Medical History",
            "Medication",
            "Extended Fam Med Hist",
            "Long Term Health Care",
            "Organ & Body Donation",
            "Medical Documentation",
        ],
    },
    {
        "name": "Final Arrangements",
        "icon": "scroll-text",
        "sections": [
            "Guardianship of Children",
            "Final Arrangements",
            "Funeral Guest List",
            "Eulogy Notes",
            "Personal Letters",
            "Estate Plans",
            "Other Contacts",
            "Funeral Receipts",
        ],
    },
]

# Placeholder sheets: no fields, just instruction text
PLACEHOLDER_SHEETS = {
    "Current Resume",
    "Transcripts",
    "Credit Report",
    "Funeral Receipts",
}

# key_value type sheets
KEY_VALUE_SHEETS = {
    "Document Location",
    "Emergency Plan",
    "Private Business",
    "Estate Plans",
}

# Friendly section names (sheet name -> display name)
SECTION_NAMES = {
    "Personal Info": "Personal Information",
    "Adoption Info": "Adoption Information",
    "Pet Information": "Pet Information",
    "Previous Addresses": "Previous Addresses",
    "Current Resume": "Current Resume",
    "Employment History": "Employment History",
    "Schooling History": "Schooling History",
    "Transcripts": "Transcripts",
    "Military History": "Military History",
    "Groups and Organizations": "Groups and Organizations",
    "Credit Report": "Credit Report",
    "Document Location": "Document Location",
    "Emergency Plan": "Emergency Plan",
    "Travel Info": "Travel Information",
    "Safes & Storage": "Safes & Storage",
    "Tax Issues": "Tax Issues",
    "Property Tax": "Property Tax",
    "Est Tax Payments": "Estimated Tax Payments",
    "Data Backup Plans": "Data Backup Plans",
    "Passwords & Logins": "Passwords & Logins",
    "Account Numbers": "Account Numbers",
    "Courses": "Courses",
    "Bank Accounts": "Bank Accounts",
    "Investments": "Investments",
    "Retirement Plans": "Retirement Plans",
    "Private Business": "Private Business",
    "Real Estate": "Real Estate",
    "Vehicles": "Vehicles",
    "Valuables Inventory": "Valuables Inventory",
    "Debts Owed to You": "Debts Owed to You",
    "Other Assets": "Other Assets",
    "Subscriptions": "Subscriptions",
    "Loan Obligations": "Loan Obligations",
    "Credit Cards": "Credit Cards",
    "Other Debts": "Other Debts",
    "Life Insurance": "Life Insurance",
    "Health Insurance": "Health Insurance",
    "Car Insurance": "Car Insurance",
    "Vehicle Insurance": "Vehicle Insurance",
    "Home-Renter Insurance": "Home-Renter Insurance",
    "Medical History": "Medical History",
    "Medication": "Medication",
    "Extended Fam Med Hist": "Extended Family Medical History",
    "Long Term Health Care": "Long Term Health Care",
    "Organ & Body Donation": "Organ & Body Donation",
    "Medical Documentation": "Medical Documentation",
    "Guardianship of Children": "Guardianship of Children",
    "Final Arrangements": "Final Arrangements",
    "Funeral Guest List": "Funeral Guest List",
    "Eulogy Notes": "Eulogy Notes",
    "Personal Letters": "Personal Letters",
    "Estate Plans": "Estate Plans",
    "Other Contacts": "Other Contacts",
    "Funeral Receipts": "Funeral Receipts",
}

# Section descriptions (from the title row or context)
SECTION_DESCRIPTIONS = {
    "Personal Info": "Basic personal details for each family member.",
    "Adoption Info": "Adoption records and birth parent information.",
    "Pet Information": "Pet details, veterinary contacts, and insurance.",
    "Previous Addresses": "History of previous residential addresses.",
    "Current Resume": "Upload copies of current resumes for all family members.",
    "Employment History": "Current and previous employment records.",
    "Schooling History": "Educational history and academic records.",
    "Transcripts": "Upload copies of college transcripts for all family members.",
    "Military History": "Military service records and details.",
    "Groups and Organizations": "Clubs, professional organizations, civic groups, etc.",
    "Credit Report": "Upload a current copy of your credit report(s). You can request free reports at annualcreditreport.com.",
    "Document Location": "Location of important documents and contacts.",
    "Emergency Plan": "Emergency meeting locations, contacts, and safety information.",
    "Travel Info": "Passport details and loyalty program information.",
    "Safes & Storage": "Safes, storage units, PO boxes, and safety deposit boxes.",
    "Tax Issues": "Tax records, CPA contacts, and tax payment history.",
    "Property Tax": "Property tax records and details.",
    "Est Tax Payments": "Estimated tax payment schedules and records.",
    "Data Backup Plans": "Data backup services, devices, and restore procedures.",
    "Passwords & Logins": "Website and service login credentials.",
    "Account Numbers": "Utility, mortgage, cable, and other account numbers.",
    "Courses": "Online courses and educational subscriptions.",
    "Bank Accounts": "Bank account details and access information.",
    "Investments": "Investments including mutual funds, annuities, and stocks.",
    "Retirement Plans": "Retirement plans including 401k, pensions, and IRAs.",
    "Private Business": "Private business details, partners, and online accounts.",
    "Real Estate": "Real estate properties, mortgages, and tax information.",
    "Vehicles": "Automobiles, motorcycles, boats, RVs, and other vehicles.",
    "Valuables Inventory": "Valuable items inventory, room by room.",
    "Debts Owed to You": "Debts and obligations owed to you.",
    "Other Assets": "Other assets such as savings bonds and stock options.",
    "Subscriptions": "Newspapers, magazines, monthly services, Netflix, etc.",
    "Loan Obligations": "Loan obligations and repayment details.",
    "Credit Cards": "Credit card details, limits, and balances.",
    "Other Debts": "Other debts and obligations.",
    "Life Insurance": "Life insurance policies including AD&D and LTD.",
    "Health Insurance": "Health insurance policies including dental and prescription drugs.",
    "Car Insurance": "Car insurance policies and coverage details.",
    "Vehicle Insurance": "Other vehicle insurance policies.",
    "Home-Renter Insurance": "Homeowner and renter insurance policies.",
    "Medical History": "Medical history for yourself and immediate family, including allergies.",
    "Medication": "Prescription medication details.",
    "Extended Fam Med Hist": "Extended family medical history.",
    "Long Term Health Care": "Long term health care directions for self and immediate family.",
    "Organ & Body Donation": "Organ, tissue, and body donation preferences.",
    "Medical Documentation": "Medical documentation locations and legal contacts.",
    "Guardianship of Children": "Guardianship designations and guardian contact information.",
    "Final Arrangements": "Final arrangement wishes and funeral details.",
    "Funeral Guest List": "List of people to notify and invite to funeral services.",
    "Eulogy Notes": "Milestones of your life to help with any eulogy.",
    "Personal Letters": "Letters to loved ones, typically opened upon the death of the writer.",
    "Estate Plans": "Estate plans, wills, trusts, and gift distributions.",
    "Other Contacts": "Other contacts to notify in case of death.",
    "Funeral Receipts": "Upload all funeral-related expenses and receipts here.",
}


def extract_fields_standard(ws, sheet_name: str) -> list[dict]:
    """
    Extract fields from a standard repeating-record sheet.
    The first record block defines the fields. We scan from row 2 until
    we see a label that already appeared (indicating the pattern is repeating).
    """
    seen_labels = []
    seen_set = set()

    # For most sheets, the first record block spans from row 2 until labels repeat.
    # We scan all rows looking for cell values, stopping when we see a duplicate label.
    max_row = ws.max_row
    max_col = ws.max_column

    for row in range(2, max_row + 1):
        row_labels = []
        for col in range(1, max_col + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                label = str(val).strip()
                if not label or label == "Table of Contents":
                    continue
                row_labels.append(label)

        # Check if any label in this row is a repeat of one we already saw
        found_repeat = False
        for label in row_labels:
            if label in seen_set:
                found_repeat = True
                break

        if found_repeat:
            break

        for label in row_labels:
            if label not in seen_set:
                seen_labels.append(label)
                seen_set.add(label)

    # Build field objects
    fields = []
    for i, label in enumerate(seen_labels, 1):
        fields.append({
            "name": label,
            "slug": make_slug(label),
            "fieldType": guess_field_type(label),
            "sortOrder": i,
        })

    return fields


def extract_fields_document_location(ws) -> list[dict]:
    """
    Document Location has a unique layout with paired columns and different
    sub-sections (birth certs, marriage, divorce, other docs, attorneys, accountants).
    Extract unique field labels across the entire sheet.
    """
    labels = []
    seen = set()
    max_row = ws.max_row
    max_col = ws.max_column

    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                label = str(val).strip()
                if not label or label == "Table of Contents":
                    continue
                if label not in seen:
                    labels.append(label)
                    seen.add(label)

    fields = []
    for i, label in enumerate(labels, 1):
        fields.append({
            "name": label,
            "slug": make_slug(label),
            "fieldType": guess_field_type(label),
            "sortOrder": i,
        })
    return fields


def extract_fields_emergency_plan(ws) -> list[dict]:
    """
    Emergency Plan is a key_value section where each row has unique labels.
    Extract all unique labels.
    """
    labels = []
    seen = set()
    max_row = ws.max_row
    max_col = ws.max_column

    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                label = str(val).strip()
                if not label or label == "Table of Contents":
                    continue
                if label not in seen:
                    labels.append(label)
                    seen.add(label)

    fields = []
    for i, label in enumerate(labels, 1):
        fields.append({
            "name": label,
            "slug": make_slug(label),
            "fieldType": guess_field_type(label),
            "sortOrder": i,
        })
    return fields


def extract_fields_private_business(ws) -> list[dict]:
    """
    Private Business is a key_value section with a complex non-repeating layout.
    Extract all unique labels from the entire sheet.
    """
    labels = []
    seen = set()
    max_row = ws.max_row
    max_col = ws.max_column

    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                label = str(val).strip()
                if not label or label == "Table of Contents":
                    continue
                if label not in seen:
                    labels.append(label)
                    seen.add(label)

    fields = []
    for i, label in enumerate(labels, 1):
        fields.append({
            "name": label,
            "slug": make_slug(label),
            "fieldType": guess_field_type(label),
            "sortOrder": i,
        })
    return fields


def extract_fields_estate_plans(ws) -> list[dict]:
    """
    Estate Plans has both a will/trust part and a gift distribution part.
    Extract all unique labels from the entire sheet.
    """
    labels = []
    seen = set()
    max_row = ws.max_row
    max_col = ws.max_column

    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                label = str(val).strip()
                if not label or label == "Table of Contents":
                    continue
                if label not in seen:
                    labels.append(label)
                    seen.add(label)

    fields = []
    for i, label in enumerate(labels, 1):
        fields.append({
            "name": label,
            "slug": make_slug(label),
            "fieldType": guess_field_type(label),
            "sortOrder": i,
        })
    return fields


def extract_fields_other_contacts(ws) -> list[dict]:
    """
    Other Contacts has sub-sections with different field layouts.
    Extract all unique field labels, skipping sub-section headers and
    static data like addresses and phone numbers.
    """
    # Sub-section headers (these are category labels, not fields per se,
    # but we include them as they identify the contact type)
    labels = []
    seen = set()
    max_row = ws.max_row
    max_col = ws.max_column

    # Known static data values to skip
    static_values = {
        "800-772-1213", "800-827-1000",
        "PO Box 105139", "P.O. Box 4500", "P.O. Box 2000",
        "Atlanta, GA 30348", "Allen, TX 75013", "Chester, PA 19022",
        "Equifax", "Experian", "TransUnion",
    }

    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                label = str(val).strip()
                if not label or label == "Table of Contents":
                    continue
                if label in static_values:
                    continue
                if label not in seen:
                    labels.append(label)
                    seen.add(label)

    fields = []
    for i, label in enumerate(labels, 1):
        fields.append({
            "name": label,
            "slug": make_slug(label),
            "fieldType": guess_field_type(label),
            "sortOrder": i,
        })
    return fields


def extract_fields_personal_letters(ws) -> list[dict]:
    """
    Personal Letters has instruction text in row 2, then repeating fields starting at row 3.
    """
    labels = []
    seen = set()
    max_col = ws.max_column

    # Fields start at row 3
    for col in range(1, max_col + 1):
        val = ws.cell(row=3, column=col).value
        if val is not None:
            label = str(val).strip()
            if label and label != "Table of Contents" and label not in seen:
                labels.append(label)
                seen.add(label)

    fields = []
    for i, label in enumerate(labels, 1):
        fields.append({
            "name": label,
            "slug": make_slug(label),
            "fieldType": guess_field_type(label),
            "sortOrder": i,
        })
    return fields


def extract_fields_medical_documentation(ws) -> list[dict]:
    """
    Medical Documentation has a repeating pattern of Document/Location/Law Firm/Contact
    with specific document type labels in between (HIPAA, Health care proxy, etc.).
    Extract the header fields plus the specific document types.
    """
    labels = []
    seen = set()
    max_row = ws.max_row
    max_col = ws.max_column

    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                label = str(val).strip()
                if not label or label == "Table of Contents":
                    continue
                if label not in seen:
                    labels.append(label)
                    seen.add(label)

    fields = []
    for i, label in enumerate(labels, 1):
        fields.append({
            "name": label,
            "slug": make_slug(label),
            "fieldType": guess_field_type(label),
            "sortOrder": i,
        })
    return fields


def extract_fields_final_arrangements(ws) -> list[dict]:
    """
    Final Arrangements has a large non-repeating block (rows 2-12) then it repeats at row 15.
    Extract unique labels from rows 2-13.
    """
    labels = []
    seen = set()
    max_col = ws.max_column

    for row in range(2, 14):  # rows 2-13
        for col in range(1, max_col + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                label = str(val).strip()
                if not label or label == "Table of Contents":
                    continue
                if label not in seen:
                    labels.append(label)
                    seen.add(label)

    # Also check for "Prepaid funeral information" label at row 28
    val = ws.cell(row=28, column=1).value
    if val:
        label = str(val).strip()
        if label and label not in seen:
            labels.append(label)
            seen.add(label)

    fields = []
    for i, label in enumerate(labels, 1):
        fields.append({
            "name": label,
            "slug": make_slug(label),
            "fieldType": guess_field_type(label),
            "sortOrder": i,
        })
    return fields


def extract_fields_home_renter_insurance(ws) -> list[dict]:
    """
    Home-Renter Insurance has a non-repeating block (rows 2-11) then repeats.
    Extract unique labels from rows 2-11.
    """
    labels = []
    seen = set()
    max_col = ws.max_column

    for row in range(2, 12):  # rows 2-11
        for col in range(1, max_col + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                label = str(val).strip()
                if not label or label == "Table of Contents":
                    continue
                if label not in seen:
                    labels.append(label)
                    seen.add(label)

    fields = []
    for i, label in enumerate(labels, 1):
        fields.append({
            "name": label,
            "slug": make_slug(label),
            "fieldType": guess_field_type(label),
            "sortOrder": i,
        })
    return fields


def extract_fields_car_insurance(ws) -> list[dict]:
    """
    Car Insurance has a record block of rows 2-4 (with some blank rows), then repeats at row 9.
    Extract from rows 2-7.
    """
    labels = []
    seen = set()
    max_col = ws.max_column

    for row in range(2, 8):
        for col in range(1, max_col + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                label = str(val).strip()
                if not label or label == "Table of Contents":
                    continue
                if label not in seen:
                    labels.append(label)
                    seen.add(label)

    fields = []
    for i, label in enumerate(labels, 1):
        fields.append({
            "name": label,
            "slug": make_slug(label),
            "fieldType": guess_field_type(label),
            "sortOrder": i,
        })
    return fields


def extract_fields_vehicle_insurance(ws) -> list[dict]:
    """
    Vehicle Insurance: same pattern as Car Insurance but with Vehicle Type added.
    Record block rows 2-7, repeats at row 9.
    """
    labels = []
    seen = set()
    max_col = ws.max_column

    for row in range(2, 8):
        for col in range(1, max_col + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                label = str(val).strip()
                if not label or label == "Table of Contents":
                    continue
                if label not in seen:
                    labels.append(label)
                    seen.add(label)

    fields = []
    for i, label in enumerate(labels, 1):
        fields.append({
            "name": label,
            "slug": make_slug(label),
            "fieldType": guess_field_type(label),
            "sortOrder": i,
        })
    return fields


def extract_fields_debts(ws) -> list[dict]:
    """
    Debts Owed to You / Other Debts: record block is rows 2-6/7, repeats at 9.
    """
    labels = []
    seen = set()
    max_col = ws.max_column

    for row in range(2, 8):
        for col in range(1, max_col + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                label = str(val).strip()
                if not label or label == "Table of Contents":
                    continue
                if label not in seen:
                    labels.append(label)
                    seen.add(label)

    fields = []
    for i, label in enumerate(labels, 1):
        fields.append({
            "name": label,
            "slug": make_slug(label),
            "fieldType": guess_field_type(label),
            "sortOrder": i,
        })
    return fields


def extract_fields_real_estate(ws) -> list[dict]:
    """
    Real Estate: record block rows 2-8, repeats at row 11.
    """
    labels = []
    seen = set()
    max_col = ws.max_column

    for row in range(2, 10):
        for col in range(1, max_col + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                label = str(val).strip()
                if not label or label == "Table of Contents":
                    continue
                if label not in seen:
                    labels.append(label)
                    seen.add(label)

    fields = []
    for i, label in enumerate(labels, 1):
        fields.append({
            "name": label,
            "slug": make_slug(label),
            "fieldType": guess_field_type(label),
            "sortOrder": i,
        })
    return fields


def extract_fields_credit_cards(ws) -> list[dict]:
    """
    Credit Cards: large record block rows 2-8, repeats at row 10.
    """
    labels = []
    seen = set()
    max_col = ws.max_column

    for row in range(2, 10):
        for col in range(1, max_col + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                label = str(val).strip()
                if not label or label == "Table of Contents":
                    continue
                if label not in seen:
                    labels.append(label)
                    seen.add(label)

    fields = []
    for i, label in enumerate(labels, 1):
        fields.append({
            "name": label,
            "slug": make_slug(label),
            "fieldType": guess_field_type(label),
            "sortOrder": i,
        })
    return fields


def extract_fields_eulogy_notes(ws) -> list[dict]:
    """
    Eulogy Notes: row 2=Who, row 4=Notes, then repeats at row 11.
    """
    labels = []
    seen = set()
    max_col = ws.max_column

    for row in range(2, 10):
        for col in range(1, max_col + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                label = str(val).strip()
                if not label or label == "Table of Contents":
                    continue
                if label not in seen:
                    labels.append(label)
                    seen.add(label)

    fields = []
    for i, label in enumerate(labels, 1):
        fields.append({
            "name": label,
            "slug": make_slug(label),
            "fieldType": guess_field_type(label),
            "sortOrder": i,
        })
    return fields


def extract_fields_guardianship(ws) -> list[dict]:
    """
    Guardianship of Children: record block rows 2-6, repeats at row 12.
    """
    labels = []
    seen = set()
    max_col = ws.max_column

    for row in range(2, 11):
        for col in range(1, max_col + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                label = str(val).strip()
                if not label or label == "Table of Contents":
                    continue
                if label not in seen:
                    labels.append(label)
                    seen.add(label)

    fields = []
    for i, label in enumerate(labels, 1):
        fields.append({
            "name": label,
            "slug": make_slug(label),
            "fieldType": guess_field_type(label),
            "sortOrder": i,
        })
    return fields


def extract_fields_travel_info(ws) -> list[dict]:
    """
    Travel Info has two sub-sections: passport fields (rows 2-8) and loyalty program
    fields (rows 10+). The passport fields repeat every 2 rows, loyalty fields repeat
    every 2 rows. Extract unique labels from rows 2 and 10.
    """
    labels = []
    seen = set()
    max_col = ws.max_column

    # Passport fields from row 2
    for col in range(1, max_col + 1):
        val = ws.cell(row=2, column=col).value
        if val is not None:
            label = str(val).strip()
            if label and label != "Table of Contents" and label not in seen:
                labels.append(label)
                seen.add(label)

    # Loyalty program fields from row 10
    for col in range(1, max_col + 1):
        val = ws.cell(row=10, column=col).value
        if val is not None:
            label = str(val).strip()
            if label and label != "Table of Contents" and label not in seen:
                labels.append(label)
                seen.add(label)

    fields = []
    for i, label in enumerate(labels, 1):
        fields.append({
            "name": label,
            "slug": make_slug(label),
            "fieldType": guess_field_type(label),
            "sortOrder": i,
        })
    return fields


def extract_fields_for_sheet(ws, sheet_name: str) -> list[dict]:
    """Route to the correct extraction method per sheet."""
    if sheet_name == "Travel Info":
        return extract_fields_travel_info(ws)
    elif sheet_name == "Document Location":
        return extract_fields_document_location(ws)
    elif sheet_name == "Emergency Plan":
        return extract_fields_emergency_plan(ws)
    elif sheet_name == "Private Business":
        return extract_fields_private_business(ws)
    elif sheet_name == "Estate Plans":
        return extract_fields_estate_plans(ws)
    elif sheet_name == "Other Contacts":
        return extract_fields_other_contacts(ws)
    elif sheet_name == "Personal Letters":
        return extract_fields_personal_letters(ws)
    elif sheet_name == "Medical Documentation":
        return extract_fields_medical_documentation(ws)
    elif sheet_name == "Final Arrangements":
        return extract_fields_final_arrangements(ws)
    elif sheet_name == "Home-Renter Insurance":
        return extract_fields_home_renter_insurance(ws)
    elif sheet_name == "Car Insurance":
        return extract_fields_car_insurance(ws)
    elif sheet_name == "Vehicle Insurance":
        return extract_fields_vehicle_insurance(ws)
    elif sheet_name in ("Debts Owed to You", "Other Debts"):
        return extract_fields_debts(ws)
    elif sheet_name == "Real Estate":
        return extract_fields_real_estate(ws)
    elif sheet_name == "Credit Cards":
        return extract_fields_credit_cards(ws)
    elif sheet_name == "Eulogy Notes":
        return extract_fields_eulogy_notes(ws)
    elif sheet_name == "Guardianship of Children":
        return extract_fields_guardianship(ws)
    else:
        return extract_fields_standard(ws, sheet_name)


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    output = {
        "version": "4.0",
        "categories": [],
    }

    for cat_idx, cat_def in enumerate(CATEGORIES, 1):
        category = {
            "name": cat_def["name"],
            "slug": make_slug(cat_def["name"]),
            "icon": cat_def["icon"],
            "sortOrder": cat_idx,
            "sections": [],
        }

        for sec_idx, sheet_name in enumerate(cat_def["sections"], 1):
            section_name = SECTION_NAMES.get(sheet_name, sheet_name)
            section_slug = make_slug(section_name)
            description = SECTION_DESCRIPTIONS.get(sheet_name, "")

            if sheet_name in PLACEHOLDER_SHEETS:
                # Placeholder section - get description from row 2
                ws = wb[sheet_name]
                row2_val = ws.cell(row=2, column=1).value
                if row2_val:
                    description = str(row2_val).strip()

                section = {
                    "name": section_name,
                    "slug": section_slug,
                    "type": "placeholder",
                    "sortOrder": sec_idx,
                    "description": description,
                }
            else:
                ws = wb[sheet_name]

                if sheet_name in KEY_VALUE_SHEETS:
                    section_type = "key_value"
                else:
                    section_type = "table"

                fields = extract_fields_for_sheet(ws, sheet_name)

                section = {
                    "name": section_name,
                    "slug": section_slug,
                    "type": section_type,
                    "sortOrder": sec_idx,
                    "description": description,
                    "fields": fields,
                }

            category["sections"].append(section)

        output["categories"].append(category)

    # Write output
    with open(OUTPUT_PATH, "w") as f:
        json.dump(output, f, indent=2)
        f.write("\n")

    print(f"Wrote {OUTPUT_PATH}")

    # Print summary
    total_sections = 0
    total_fields = 0
    for cat in output["categories"]:
        for sec in cat["sections"]:
            total_sections += 1
            if "fields" in sec:
                total_fields += len(sec["fields"])

    print(f"Categories: {len(output['categories'])}")
    print(f"Sections: {total_sections}")
    print(f"Total fields: {total_fields}")


if __name__ == "__main__":
    main()
