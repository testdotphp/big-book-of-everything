#!/usr/bin/env python3
"""Import Big Book of Everything spreadsheet into SQLite database."""

import sqlite3
import sys
import re
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Install openpyxl: pip install openpyxl")
    sys.exit(1)

CATEGORY_ICONS = {
    "Personal Info": "user",
    "Finances": "dollar-sign",
    "Insurance": "shield",
    "Medical": "heart-pulse",
    "Home & Property": "home",
    "Legal & Estate": "scroll-text",
    "Digital": "monitor",
    "Education": "book-open",
    "Employment": "clipboard",
    "Misc": "layout",
}


def slugify(text: str) -> str:
    return re.sub(r"(^-|-$)", "", re.sub(r"[^a-z0-9]+", "-", text.lower()))


def detect_type(sheet) -> str:
    """Heuristic: if first row looks like headers and there are multiple data rows, it's a table."""
    rows = list(sheet.iter_rows(min_row=1, max_row=min(20, sheet.max_row or 1), values_only=True))
    if len(rows) < 2:
        return "key_value"

    first_row = rows[0]
    if first_row and all(isinstance(c, str) for c in first_row if c is not None):
        data_rows = [r for r in rows[1:] if any(c is not None for c in r)]
        if len(data_rows) > 1:
            return "table"

    return "key_value"


def import_sheet(conn, sheet, section_id, section_type):
    """Import data from a sheet into the database."""
    cursor = conn.cursor()
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        return

    if section_type == "table":
        headers = [str(h).strip() if h else f"Column {i+1}" for i, h in enumerate(rows[0])]
        col_indices = [i for i, h in enumerate(rows[0]) if h is not None]

        field_ids = {}
        for order, col_idx in enumerate(col_indices):
            name = headers[col_idx]
            slug = slugify(name)
            if not slug:
                slug = f"col-{order}"
            cursor.execute(
                "INSERT INTO fields (section_id, name, slug, field_type, sort_order) VALUES (?, ?, ?, 'text', ?)",
                (section_id, name, slug, order),
            )
            field_ids[col_idx] = cursor.lastrowid

        for row_order, row in enumerate(rows[1:]):
            if not any(row[i] for i in col_indices if i < len(row)):
                continue
            cursor.execute(
                "INSERT INTO records (section_id, sort_order) VALUES (?, ?)",
                (section_id, row_order),
            )
            record_id = cursor.lastrowid
            for col_idx in col_indices:
                val = row[col_idx] if col_idx < len(row) else None
                if val is not None:
                    cursor.execute(
                        'INSERT INTO "values" (field_id, record_id, value) VALUES (?, ?, ?)',
                        (field_ids[col_idx], record_id, str(val)),
                    )
    else:
        for order, row in enumerate(rows):
            if not row or row[0] is None:
                continue
            label = str(row[0]).strip()
            if not label:
                continue
            slug = slugify(label)
            if not slug:
                slug = f"field-{order}"

            cursor.execute(
                "INSERT INTO fields (section_id, name, slug, field_type, sort_order) VALUES (?, ?, ?, 'text', ?)",
                (section_id, label, slug, order),
            )
            field_id = cursor.lastrowid

            value = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            if value:
                cursor.execute(
                    'INSERT INTO "values" (field_id, record_id, value) VALUES (?, NULL, ?)',
                    (field_id, value),
                )

    conn.commit()


def main():
    if len(sys.argv) < 3:
        print(f"Usage: {sys.argv[0]} <spreadsheet.xlsx> <output.db>")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    db_path = sys.argv[2]

    wb = load_workbook(xlsx_path, data_only=True)

    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")

    conn.executescript('''
        CREATE TABLE IF NOT EXISTS categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            slug TEXT NOT NULL UNIQUE,
            icon TEXT DEFAULT 'folder',
            sort_order INTEGER NOT NULL DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS sections (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            category_id INTEGER NOT NULL REFERENCES categories(id) ON DELETE CASCADE,
            name TEXT NOT NULL,
            slug TEXT NOT NULL,
            type TEXT NOT NULL DEFAULT 'key_value',
            sort_order INTEGER NOT NULL DEFAULT 0,
            UNIQUE(category_id, slug)
        );
        CREATE TABLE IF NOT EXISTS fields (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            section_id INTEGER NOT NULL REFERENCES sections(id) ON DELETE CASCADE,
            name TEXT NOT NULL,
            slug TEXT NOT NULL,
            field_type TEXT NOT NULL DEFAULT 'text',
            sort_order INTEGER NOT NULL DEFAULT 0,
            UNIQUE(section_id, slug)
        );
        CREATE TABLE IF NOT EXISTS records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            section_id INTEGER NOT NULL REFERENCES sections(id) ON DELETE CASCADE,
            sort_order INTEGER NOT NULL DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS "values" (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            field_id INTEGER NOT NULL REFERENCES fields(id) ON DELETE CASCADE,
            record_id INTEGER REFERENCES records(id) ON DELETE CASCADE,
            value TEXT
        );
    ''')

    print(f"Found {len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)}")

    keyword_map = {
        "Personal Info": ["personal", "family", "emergency", "contact", "member"],
        "Finances": ["bank", "credit", "invest", "retire", "debt", "budget", "financial", "account", "loan", "mortgage"],
        "Insurance": ["insurance", "policy", "coverage"],
        "Medical": ["medical", "doctor", "medication", "allerg", "health", "immuniz", "pharmacy", "dental", "vision"],
        "Home & Property": ["home", "house", "property", "maintenance", "warranty", "appliance", "vehicle", "auto", "car"],
        "Legal & Estate": ["will", "trust", "estate", "legal", "attorney", "power of attorney", "beneficiar", "document"],
        "Digital": ["online", "digital", "subscription", "software", "license", "password", "website", "email"],
        "Education": ["school", "education", "transcript", "certif", "degree", "college"],
        "Employment": ["employ", "job", "work", "career", "benefit", "salary", "resume"],
    }

    cursor = conn.cursor()
    cat_ids = {}
    all_categories = list(CATEGORY_ICONS.keys())
    for order, cat_name in enumerate(all_categories):
        icon = CATEGORY_ICONS.get(cat_name, "folder")
        cursor.execute(
            "INSERT INTO categories (name, slug, icon, sort_order) VALUES (?, ?, ?, ?)",
            (cat_name, slugify(cat_name), icon, order),
        )
        cat_ids[cat_name] = cursor.lastrowid
    conn.commit()

    for sheet_order, sheet_name in enumerate(wb.sheetnames):
        sheet = wb[sheet_name]
        if sheet.max_row is None or sheet.max_row < 1:
            continue

        assigned_cat = "Misc"
        sheet_lower = sheet_name.lower()
        for cat_name, keywords in keyword_map.items():
            if any(kw in sheet_lower for kw in keywords):
                assigned_cat = cat_name
                break

        section_type = detect_type(sheet)

        cursor.execute(
            "INSERT INTO sections (category_id, name, slug, type, sort_order) VALUES (?, ?, ?, ?, ?)",
            (cat_ids[assigned_cat], sheet_name, slugify(sheet_name), section_type, sheet_order),
        )
        section_id = cursor.lastrowid

        import_sheet(conn, sheet, section_id, section_type)
        print(f"  [{assigned_cat}] {sheet_name} ({section_type}, {sheet.max_row} rows)")

    conn.close()
    print(f"\nDone. Database written to {db_path}")


if __name__ == "__main__":
    main()
